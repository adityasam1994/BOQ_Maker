from django.shortcuts import render, HttpResponse, HttpResponseRedirect, reverse
from .models import master_db, boq, master_item, master_item_edit, item, child_master, myuser, excel_boq_data, resource_category
from django.views.decorators.csrf import csrf_exempt
import json
import pickle
from openpyxl import load_workbook
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.styles.borders import Border, Side
#from tkinter import *
#from tkinter.filedialog import asksaveasfile

# Create your views here.

# win= Tk()
# win.geometry("750x250")

def home(request):
    if 'username' in request.session:
        print(request.session['username'])
        return render(request, 'home.html')
    else:
        return HttpResponseRedirect(reverse("loginpage"))

def loginpage(request):
    if 'username' in request.session:
        return HttpResponseRedirect(reverse("home"))
    else:
        return render(request, "loginpage.html")

@csrf_exempt
def loginuser(request):
    data = request.POST.get("data")
    data_split = data.split("/////")
    username = data_split[0]
    password = data_split[1]
    
    mu = myuser.objects.filter(username = username, password=password, active=True)
    data = "error"
    if len(mu) != 0:
        request.session['username'] = mu[0].username
    
        data = "ok"

    response_data = {}
    response_data['data'] = data
    return HttpResponse(
        json.dumps(response_data),
        content_type="application/json"
    )

@csrf_exempt
def logoutuser(request):
    del request.session['username']

    response_data = {}
    response_data['data'] = 'ok'
    return HttpResponse(
        json.dumps(response_data),
        content_type="application/json"
    )

def create_boq(request):
    if request.method == "POST":
        name = request.POST.get("boqname")
        code = request.POST.get("boqcode")
        desc = request.POST.get("boqdesc")

        bo = boq()
        bo.name = name
        bo.code = code
        bo.description = desc
        bo.save()
        return HttpResponseRedirect(reverse('load_excel_boq', args=(bo.id,)))
    else:
        return HttpResponseRedirect(reverse('home'))

def open_excel_boq(request, bid):
    return render(request, "excel_boq.html", {'bid':bid})

def read_excel_file(wb_obj, heading_num, desc_num, quantity_num, uom_num, first_number):
    sheet_obj = wb_obj.active
    row_num = sheet_obj.max_row
    col_num = sheet_obj.max_column

    data = ""
    for i in range(int(first_number), row_num+1):
        head_obj = sheet_obj.cell(row = i, column = int(heading_num)).value
        desc_obj = sheet_obj.cell(row = i, column = int(desc_num)).value
        qt_obj = sheet_obj.cell(row = i, column = int(quantity_num)).value
        uom_obj = sheet_obj.cell(row = i, column = int(uom_num)).value

        if len(str(desc_obj)) > 0:
            if data == "":
                data = str(head_obj) + "^^^^^" + str(desc_obj) + "^^^^^" + str(qt_obj) + "^^^^^" + str(uom_obj) + "^^^^^None^^^^^f1^^^^^None"
            else:
                data = data + "/////" + str(head_obj) + "^^^^^" + str(desc_obj) + "^^^^^" + str(qt_obj) + "^^^^^" + str(uom_obj) + "^^^^^None^^^^^f1^^^^^None"
    return data

def excel_boq(request, bid):
    if request.method == "POST":
        wb = load_workbook(filename=request.FILES['uploaded_file'].file)
        heading_num = request.POST.get("heading_number")
        desc_num = request.POST.get("desc_number")
        quantity_num = request.POST.get("quantity_number")
        uom_num = request.POST.get("uom_number")
        first_number = request.POST.get("first_number")

        data = read_excel_file(wb, heading_num, desc_num, quantity_num, uom_num, first_number)
        bq = excel_boq_data()
        bq.boq_id = bid
        bq.data = data
        bq.save()

        return HttpResponseRedirect(reverse("show_boq", args=(bid, bq.id)))
    else:
        return HttpResponseRedirect(reverse("show_boq", args=(bid,)))

def show_boq(request, bid, bq = 9999999):
    bo = boq.objects.get(id=bid)
    data = None
    if bq != 9999999:
        data = excel_boq_data.objects.get(id=bq)

    return render(request, "show_boq.html", {'boq':bo, 'excel_data':data})

    # New Methods

def load_excel_boq(request, boqid):
    if request.method == "POST": # Can be posted from home page while making new boq or from the BOQ page itself when uploading excel
        bq = boq.objects.get(id=boqid)
        wb = load_workbook(filename=request.FILES['uploaded_file'].file)
        heading_num = request.POST.get("heading_number")
        desc_num = request.POST.get("desc_number")
        quantity_num = request.POST.get("quantity_number")
        uom_num = request.POST.get("uom_number")
        first_number = request.POST.get("first_number")

        data = read_excel_file(wb, heading_num, desc_num, quantity_num, uom_num, first_number)
        bq.boq_id = boqid
        bq.data = data
        bq.save()
        
        return HttpResponseRedirect(reverse("show_excel_boq", args=(boqid,)))

    else:
        excel = boq.objects.get(id=boqid)
        return HttpResponseRedirect(reverse("show_excel_boq", args=(boqid,)))

def show_excel_boq(request, bo):
    if "username" in request.session:
        if bo == "select":
            return HttpResponseRedirect(reverse("home"))
        else:
            bq = boq.objects.get(id=bo)
            return render(request, "myboq.html", {'boq':bq})
    else:
        return HttpResponseRedirect(reverse("loginpage"))
# Method to load all the resources as a string

@csrf_exempt
def fetch_resources(request):
    resources = master_db.objects.all()
    big_data = ""
    for resource in resources:
        if big_data == "":
                       # Resource ID             Resource code               Resource Name              Resource Price
            big_data = str(resource.id) + ",," + str(resource.code) + ",," + str(resource.name) +",," + str(resource.price)
        else:
            big_data = big_data + "//" + str(resource.id) + ",," + str(resource.code) + ",," + str(resource.name) +",," + str(resource.price)

    response_data = {}
    response_data['big_data'] = big_data
    return HttpResponse(
        json.dumps(response_data),
        content_type="application/json"
    )

@csrf_exempt
def reload_masters(request):
    resources = master_item.objects.all()
    big_data = ""
    for resource in resources:
        if big_data == "":
                       # Resource ID             Resource code               Resource Name     
            big_data = str(resource.id) + ",," + str(resource.code) + ",," + str(resource.name)
        else:
            big_data = big_data + "//" + str(resource.id) + ",," + str(resource.code) + ",," + str(resource.name)

    response_data = {}
    response_data['big_data'] = big_data
    return HttpResponse(
        json.dumps(response_data),
        content_type="application/json"
    )

@csrf_exempt
def reload_resources(request):
    resources = master_db.objects.all()
    big_data = ""
    for resource in resources:
        if big_data == "":
                       # Resource ID             Resource code               Resource Name     
            big_data = str(resource.id) + ",," + str(resource.code) + ",," + str(resource.name)
        else:
            big_data = big_data + "//" + str(resource.id) + ",," + str(resource.code) + ",," + str(resource.name)

    response_data = {}
    response_data['big_data'] = big_data
    return HttpResponse(
        json.dumps(response_data),
        content_type="application/json"
    )

# Method to get the price of a perticular resource

@csrf_exempt
def get_resource_price(request):
    resource_id = request.POST.get("data")
    resource = master_db.objects.filter(id = int(resource_id))
    price = 0
    if len(resource) > 0:
        price = resource[0].price

    response_data = {}
    response_data['price'] = price
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

#Saving the master as new

@csrf_exempt
def save_as_new_master(request):
    data = request.POST.get("data")

    split1 = data.split("___")
    boqid = split1[0]
    mid = split1[1]
    code = split1[2]
    name = split1[3]
    remark = split1[4]
    resources = split1[5]

    master = master_item()
    master.code = code
    master.name = name
    master.remark = remark
    master.components = resources
    master.save()

    response_data = {}
    response_data['master_id'] = master.id
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

# Saving the master from the modal

@csrf_exempt
def save_master(request):
    data = request.POST.get("data")

    split1 = data.split("___")
    boqid = split1[0]
    mid = split1[1]
    code = split1[2]
    name = split1[3]
    remark = split1[4]
    resources = split1[5]

    status = "okay"
    if mid != "999999":
        if boqid == "999999":
            master = master_item.objects.filter(id = int(mid))
            if len(master) == 0:
                master = master_item()
            else:
                master = master[0]
            master_check = master_item.objects.filter(code = code)
            if len(master_check) == 0:
                master.code = code
                master.name = name
                master.remark = remark
                master.components = resources
                master.save()
            else:
                status = "exist"
        else:
            master = master_item_edit.objects.filter(master_id=int(mid), boq_id = int(boqid))
            if len(master) > 0:
                master = master[0]
            else:
                master = master_item_edit()

            master.boq_id = boqid
            master.master_id = mid
            master.code = code
            master.name = name
            master.remark = remark
            master.components = resources
            master.save()

    else:
        master = master_item.objects.filter(id = int(mid))
        if len(master) == 0:
            master = master_item()
        else:
            master = master[0]
        master_check = master_item.objects.filter(code = code)
        if len(master_check) == 0:
            master.code = code
            master.name = name
            master.remark = remark
            master.components = resources
            master.save()
        else:
            status = "exist"

    response_data = {}
    response_data['data'] = str(master.id) + "/" + status
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

# Load and send all the master items

@csrf_exempt
def load_options(request):
    master_items = master_item.objects.all()

    big_data = ""
    for item in master_items:
        if big_data == "":
            big_data = str(item.id) + "___" + str(item.code) + "___" + str(item.name) + "___" + str(item.remark)
        else:
            big_data = big_data + "//" + str(item.id) + "___" + str(item.code) + "___" + str(item.name) + "___" + str(item.remark)

    response_data = {}
    response_data['bigdata'] = big_data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

# Method to calculate price of a master

def calculate_master_price(resources, tp):
    ress = resources.split("//") # Splitting to individual resources
    total = 0
    for res in ress:
        res_split = res.split(",,") # Splitting to components
        if tp == 'org':
            r = master_db.objects.filter(id = int(res_split[0]))
            if len(r) > 0:
                total = total + float(res_split[1])*r.price
        else:
            total = total + float(res_split[-1])

    return total

# Getting price of the master element

@csrf_exempt
def get_master_price(request):
    master_id = request.POST.get("data").split("/")[1]
    boqid = request.POST.get("data").split("/")[0]
    tp = 'edited'
    master = master_item_edit.objects.filter(master_id = int(master_id), boq_id = int(boqid))
    if len(master) == 0:
        tp = 'org'
        master = master_item.objects.get(id=int(master_id))
    else:
        master = master[0]

    response_data = {}
    response_data['price'] = calculate_master_price(master.components, tp)
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

# Getting the details of master element to edit

@csrf_exempt
def get_master_details(request):
    master_id = request.POST.get("data").split("/")[1]
    boq_id = request.POST.get("data").split("/")[0]
    
    master = master_item_edit.objects.filter(boq_id = int(boq_id), master_id = int(master_id))
    if len(master) == 0:
        master = master_item.objects.get(id=int(master_id))
    else:
        master = master[0]
    data = master.code +"___" + master.name + "___" + master.remark + "___" + master.components

    response_data = {}
    response_data['data'] = data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

# Getting master item detail for master page

@csrf_exempt
def get_master_details2(request):
    master_id = request.POST.get("data")
    master = master_item.objects.get(id = int(master_id))
    
    data = master.code +"___" + master.name + "___" + master.remark + "___" + master.components

    response_data = {}
    response_data['data'] = data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def get_category_name(request):
    data = request.POST.get("data")
    category = resource_category.objects.filter(category_value = data)
    nm = data
    if len(category) > 0:
        nm = category[0].category_name

    print(nm)
    
    response_data = {}
    response_data['data'] = nm
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

def clear_items(boqid):
    category = item.objects.filter(boq_id = boqid)
    item_ids = []
    l = len(category)
    for i in range(l):
        item_ids.append(category[i].id)
        category[i].delete()

    return item_ids

@csrf_exempt
def save_boq(request):
    data = request.POST.get("data")
    split1 = data.split(";;;;;")
    boqid = split1[0]
    factors = split1[1]
    base_data = split1[2]

    factors_split = factors.split("/")
    myboq = boq.objects.get(id = int(boqid))
    myboq.factor1 = factors_split[0]
    myboq.factor2 = factors_split[1]
    myboq.factor3 = factors_split[2]
    myboq.factor4 = factors_split[3]

    myboq.save()

    masters = base_data.split("--------")
    item_ids = clear_items(boqid)

    for master in masters:
        master_split = master.split("_____")
        heading = master_split[0]
        dry = master_split[1]
        sell = master_split[2]
        itms = master_split[3]

        category = item()
        category.heading = heading
        category.boq_id = boqid
        category.dry_amount = dry.replace(",","")
        category.sell_amount = sell.replace(",","")

        category.save()

        itmssp = itms.split(">>>>>")
        for itm in itmssp:
            items_split = itm.split("........")
            value = items_split[0]
            quantity = items_split[1]
            uom = items_split[2]
            price = items_split[3].replace(",","")
            amount = items_split[4].replace(",","")
            fact = items_split[5]
            sprice = items_split[6].replace(",","")
            samount = items_split[7].replace(",","")
            remark = items_split[8]

            if value != "select":
                for it in item_ids:
                    childs = child_master.objects.filter(item_id = it)
                    l = len(childs)
                    for i in range(l):
                        childs[i].delete()

                childs = child_master()
                childs.item_id = category.id
                childs.master_id = value
                childs.quantity = quantity
                childs.uom = uom
                childs.price = price.replace(",","")
                childs.amount = amount.replace(",","")
                childs.factor = fact
                childs.sell_price = sprice.replace(",","")
                childs.sell_amount = samount.replace(",","")
                childs.remark = remark

                childs.save()

    response_data = {}
    response_data['data'] = 'test'
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

def edit_boq(request, bid):
    bo = boq.objects.get(id=bid)
    return render(request, "edit_boq.html", {'boq':bo})

@csrf_exempt
def load_blocks(request):
    bid = request.POST.get("data")
    myitems = item.objects.filter(boq_id = int(bid))
    data = ""
    for it in myitems:
        if data == "":
            data = str(it.heading) + "/////" + str(it.dry_amount) + "/////" + str(it.sell_amount) + "/////" + str(it.id)
        else:
            data = data + ">>>>>" + str(it.heading) + "/////" + str(it.dry_amount) + "/////" + str(it.sell_amount) + "/////" + str(it.id)
    
    response_data = {}
    response_data['data'] = data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def get_masters(request):
    data = request.POST.get("data")
    bid = data.split("/")[0]
    item_id = data.split("/")[1]

    masters = child_master.objects.filter(item_id = int(item_id))
    bigdata = ""
    for master in masters:
        if bigdata == "":
            bigdata = str(master.master_id) + "/////" + str(master.quantity) + "/////" + str(master.price) + "/////" + str(master.amount) + \
            "/////" + str(master.sell_price) + "/////" + str(master.sell_amount) + "/////" + str(master.remark) + "/////" + str(master.factor) + "/////" + str(master.uom)

        else:
            bigdata = bigdata + ">>>>>" + str(master.master_id) + "/////" + str(master.quantity) + "/////" + str(master.price) + "/////" + str(master.amount) + \
            "/////" + str(master.sell_price) + "/////" + str(master.sell_amount) + "/////" + str(master.remark) + "/////" + str(master.factor) + "/////" + str(master.uom)

    response_data = {}
    response_data['data'] = bigdata
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )


def all_resources(request):
    if "username" in request.session:
        return render(request, "new_resources.html")
    else:
        return HttpResponseRedirect(reverse("loginpage"))

@csrf_exempt
def update_resource(request):
    alldata = request.POST.get('data')
    data_split = alldata.split("/////")
    res = master_db.objects.get(id = int(data_split[0]))
    
    res.code = data_split[1]
    res.name = data_split[2]
    res.price = data_split[3]
    res.category = data_split[4]
    res.remark = data_split[5]
    res.uom = data_split[6]
    res.save()

    response_data = {}
    response_data['data'] = 'test'
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

def all_masters(request):
    if "username" in request.session:
        return render(request, "masters.html")
    else:
        return HttpResponseRedirect(reverse("loginpage"))

@csrf_exempt
def create_new_resource(request):
    data = request.POST.get("data")
    data_split = data.split("/////")
    code = data_split[0]
    name = data_split[1]
    price = data_split[2]
    category = data_split[3]
    remark  = data_split[4]

    res = master_db()
    res.code = code
    res.name = name
    res.price = price
    res.category = category
    res.remark = remark

    res.save()

    response_data = {}
    response_data['data'] = str(res.id)
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def delete_resource_db(request):
    rid = request.POST.get("data")
    res = master_db.objects.get(id = int(rid))
    res.delete()

    response_data = {}
    response_data['data'] = rid
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

def test(request, boq_id):
    return render(request, "test.html")

def calculate_master_price_new(comps, tp):
    csplit = comps.split("//")
    total = 0
    for cs in csplit:
        split2 = cs.split(",,")
        if tp == "org":
            res = master_db.objects.filter(id = split2[0])
            if len(res) > 0:
                total += float(res[0].price)
        else:
            res = split2[3]
            total += float(res)

    return total

@csrf_exempt
def get_master_detail_new(request):
    data = request.POST.get("data").split("/////")
    boq_id = data[0]
    master_id = data[1]

    tp = "edit"
    master = master_item_edit.objects.filter(boq_id = boq_id, master_id = master_id)
    if len(master) == 0:
        tp = "org"
        master = master_item.objects.filter(id = master_id)[0]
    else:
        master = master[0]

    t_prc = calculate_master_price_new(master.components, tp)

    master_data = master.code +"___" + master.name + "___" + master.remark + "___" + master.components + "___" + str(t_prc) + "___" + tp

    response_data = {}
    response_data['master_data'] = master_data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def get_my_master_details(request):
    mid = request.POST.get("mid")
    master = master_item.objects.get(id = mid)

    t_prc = calculate_master_price_new(master.components, "org")

    master_data = master.code +"___" + master.name + "___" + master.remark + "___" + master.components + "___" + str(t_prc)

    response_data = {}
    response_data['master_data'] = master_data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def get_resource_detail_new(request):
    data = request.POST.get("data")

    resource = master_db.objects.filter(id = int(data))

    if len(resource) > 0:
        resource= resource[0]
        resource_data = str(resource.code) + "/////" + str(resource.name) + "/////" + str(resource.price) + "/////" + str(resource.category) + "/////" + str(resource.remark) + "/////" + str(resource.uom)
    else:
        resource_data = "None" + "/////" + "None" + "/////" + "0" + "/////" + "None" + "/////" + "None"

    response_data = {}
    response_data['resource_data'] = resource_data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )
    
@csrf_exempt
def get_child_master_detail(request):
    data = request.POST.get("data").split("/////")
    item_id =  data[0]
    master_id = data[1]

    child = child_master.objects.get(item_id = item_id, master_id = master_id)

    child_data = str(child.item_id) + "/////" + str(child.master_id) + "/////" + str(child.quantity) + "/////" + str(child.uom) + "/////" + str(child.price) + "/////" + str(child.amount) + "/////" + str(child.factor) + "/////" + str(child.sell_price) + "/////" + str(child.sell_amount) + "/////" + str(child.remark)

    response_data = {}
    response_data['child_data'] = child_data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def get_excel_boq_data(request):
    data = request.POST.get("data")

    excel = excel_boq_data.objects.filter(boq_id = data)

    excel_data=""
    if len(excel) > 0:
        excel_data = str(excel[0].boq_id) + "########" + str(excel[0].data)

    response_data = {}
    response_data['excel_data'] = excel_data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def get_boq_data(request):
    data = request.POST.get("data")

    bq = boq.objects.get(id = int(data))

    boq_data = str(bq.code) + "/////" + str(bq.name) + "/////" + str(bq.description) + "/////" + str(bq.factor1) + "/////" + str(bq.factor2) + "/////" + str(bq.factor3) + "/////" + str(bq.factor4)

    response_data = {}
    response_data['boq_data'] = boq_data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def get_boq_headings(request):
    data = request.POST.get("data")
    boq_heading = item.objects.filter(boq_id = int(data))

    heading_data = str(boq_heading[0].heading) + "/////" + str(boq_heading[0].boq_id)

    response_data = {}
    response_data['heading_data'] = heading_data
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def update_master(request):
    data = request.POST.get("data")
    data_split = data.split("^^^^^")
    mdata = data_split[0].split("/////")
    rdata = data_split[1]

    master = master_item.objects.get(id = mdata[0])
    master.code = mdata[1]
    master.name = mdata[2]
    master.remark = mdata[3]
    master.components = rdata
    master.save()

    response_data = {}
    response_data['data'] = "done"
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def save_local_master(request):
    data = request.POST.get("data")
    data_split = data.split("^^^^^")
    mdata = data_split[0].split("/////")
    rdata = data_split[1]

    
    master = master_item()
    master.code = mdata[0]
    master.name = mdata[1]
    master.remark = mdata[2]
    master.components = rdata
    master.save()

    response_data = {}
    response_data['data'] = master.id
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )


@csrf_exempt
def save_excel_master(request):
    data = request.POST.get("data")
    sp1 = data.split("^^^^^")
    sp = sp1[0].split("/////")
    master = master_item_edit.objects.filter(boq_id = int(sp[0]), master_id = int(sp[1]))
    if len(master) > 0:
        master = master[0]
    else:
        master = master_item_edit()
    
    master.boq_id = sp[0]
    master.master_id = sp[1]
    master.code = sp[2]
    master.name = sp[3]
    master.remark = sp[4]
    master.components = sp1[1]
    master.save()

    response_data = {}
    response_data['data'] = "done"
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def save_add_excel_master(request):
    data = request.POST.get("data")
    sp1 = data.split("^^^^^")
    sp = sp1[0].split("/////")
    
    master = master_item()
    
    master.code = sp[0]
    master.name = sp[1]
    master.remark = sp[2]
    master.components = sp1[1]
    master.save()

    response_data = {}
    response_data['data'] = master.id
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def save_my_boq(request):
    data = request.POST.get("data").split("..........")
    boqid = data[0]
    mydata = data[1]

    excel_boq = boq.objects.get(id = boqid)
    excel_boq.data = mydata
    excel_boq.save()

    with open("projects/" + excel_boq.name + ".pkl", "wb") as f:
        pickle.dump(data, f)

    response_data = {}
    response_data['status'] = "done"
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )


@csrf_exempt
def export_boq(request):
    data = request.POST.get("data").split("/")
    boqid = data[0]
    bo = boq.objects.get(id=boqid)
    boq_data = bo.data

    desc_value = data[1]
    qt_value = data[2]
    uom_value = data[3]
    price_value = data[4]
    amount_value = data[5]
    sell_price_value = data[6]
    sell_amount_value = data[7]
    rem_value = data[8]
    breakdown_value = data[9]
    resource_value = data[10]
    category_value = data[11]
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "BOQ"

    head_color = PatternFill(start_color='FF42c5f5',end_color='FF42c5f5',fill_type='solid')
    desc_color = PatternFill(start_color='FFcfdeb8',end_color='FFcfdeb8',fill_type='solid')
    res_color = PatternFill(start_color='FFffd1b5',end_color='FFffd1b5',fill_type='solid')
    total_color = PatternFill(start_color='FF03c700',end_color='FF03c700',fill_type='solid')
    main_color = PatternFill(start_color='FFe3f542',end_color='FFe3f542',fill_type='solid')

    column_num = 1

    sheet.cell(row = 2, column = 1).value = "#"
    if desc_value == "1":
        column_num += 1
        sheet.cell(row = 2, column = column_num).value = "Description"
    
    if qt_value == "1":
        column_num += 1
        sheet.cell(row = 2, column = column_num).value = "Quantity"

    if uom_value == "1":
        column_num += 1
        sheet.cell(row = 2, column = column_num).value = "UOM"

    if price_value == "1":
        column_num += 1
        sheet.cell(row = 2, column = column_num).value = " Dry Price"
    
    if amount_value == "1":
        column_num += 1
        sheet.cell(row = 2, column = column_num).value = "Dry Amount"

    if sell_price_value == "1":
        column_num += 1
        sheet.cell(row = 2, column = column_num).value = "Selling Price"
    
    if sell_amount_value == "1":
        column_num += 1
        sheet.cell(row = 2, column = column_num).value = "Selling Amount"

    if rem_value == "1":
        column_num += 1
        sheet.cell(row = 2, column = column_num).value = "Remark"

    for i in range(1,column_num+1):
        sheet.cell(row = 2, column = i).fill = head_color
        sheet.cell(row = 2, column = i).font = Font(bold=True, size = 11)

    
    thin_border = Border(left=Side(style='thin', color='ffb3b3b3'), 
                     right=Side(style='thin', color='ffb3b3b3'), 
                     top=Side(style='thin', color='ffb3b3b3'), 
                     bottom=Side(style='thin', color='ffb3b3b3'))

    spl1 = boq_data.split("/////")
    head_num = 0
    sub_head_num = 0
    ind = 2
    grand_total_dry = 0
    grand_total_sell = 0
    resource_dict = {}
    resource_count = {}
    category_sum = {}

    dry_col = 1
    sell_col = 1

    for sp in spl1:
        ind += 1
        spl2 = sp.split("^^^^^")
        num = spl2[0]
        desc = spl2[1]
        qt = spl2[2]
        uom = spl2[3]
        mid = spl2[4]
        fct = spl2[5]
        rem = spl2[6]

        if num == "None":
            num=""
        if desc == "None":
            desc=""
        if qt == "None":
            qt=""
        if uom == "None":
            uom=""
        if fct == "None":
            fct=""
        if rem == "None":
            rem=""
        if mid == "None":
            mid=""

        ff = 1
        if fct == "f1":
            ff = bo.factor1
        if fct == "f2":
            ff = bo.factor2
        if fct == "f3":
            ff = bo.factor3
        if fct == "f4":
            ff = bo.factor4

        column_num = 1

        prc=""
        if mid != "":
            prc = get_final_master_price(mid, boqid)
        else:
            prc=""

        sheet.cell(row = ind, column = 1).value = num
        if uom != "" or qt != "":
            sub_head_num += 1
            sheet.cell(row = ind, column = 1).value = str(head_num)+"."+str(sub_head_num)
            for i in range(2,column_num+1):
                sheet.cell(row = ind, column = i).fill = desc_color

        if desc_value == "1":
            column_num += 1
            sheet.cell(row = ind, column = column_num).value = desc
            column = str(chr(64 + column_num))
            sheet.column_dimensions[column].width = 50
            sheet.cell(row = ind, column = column_num).alignment = Alignment(wrap_text=True)

            if num != "":
                head_num = num
                sub_head_num=0
                sheet.cell(row = ind, column = 1).font = Font(bold=True, size = 11)
                sheet.cell(row = ind, column = column_num).font = Font(bold=True, size = 11)

            if qt == "" and uom == "":
                sheet.cell(row = ind, column = column_num).font = Font(bold=True)
                if num != "":
                    sheet.cell(row = ind, column = column_num).font = Font(bold=True, size = 14)
                    sheet.cell(row = ind, column = column_num).fill = main_color

        if qt_value == "1":
            column_num += 1
            sheet.cell(row = ind, column = column_num).value = qt
            sheet.cell(row = ind, column = column_num).number_format = numbers.FORMAT_NUMBER

        if uom_value == "1":
            column_num += 1
            sheet.cell(row = ind, column = column_num).value = uom
        
        if price_value == "1":
            column_num += 1
            if prc != "":
                sheet.cell(row = ind, column = column_num).value = (float(prc))
                sheet.cell(row = ind, column = column_num).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        if amount_value == "1":
            column_num += 1
            if qt != "" and prc != "":
                dry_col = column_num
                sheet.cell(row = ind, column = column_num).value = (prc * float(qt))
                sheet.cell(row = ind, column = column_num).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        
        if qt != "" and prc != "":
            grand_total_dry += prc * float(qt)
        
        if sell_price_value == "1":
            column_num += 1
            if prc != "":
                sheet.cell(row = ind, column = column_num).value = (prc * ff)
                sheet.cell(row = ind, column = column_num).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        if sell_amount_value == "1":
            column_num += 1
            if qt != "" and prc != "":
                sell_col = column_num
                sheet.cell(row = ind, column = column_num).value = (prc * float(qt) * ff)
                sheet.cell(row = ind, column = column_num).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                
        if qt != "" and prc != "":
            grand_total_sell += prc * float(qt) * ff

        if rem_value == "1":
            column_num += 1
            sheet.cell(row = ind, column = column_num).value = rem
        
        if uom != "" or qt != "":
            for i in range(1,column_num+1):
                sheet.cell(row = ind, column = i).fill = desc_color

        if mid != "":
            master = master_item_edit.objects.filter(master_id=mid, boq_id = boqid)
            tp="edit"
            if len(master) == 0:
                tp="org"
                master = master_item.objects.get(id=mid)
            else:
                master=master[0]

            comps = master.components
            sp1 = comps.split("//")

            if tp == "edit":
                total = 0
                for s in sp1:
                    ind += 1
                    sp2 = s.split(",,")
                    qnt = sp2[1]
                    pprc = sp2[2]
                    amnt = sp2[3]
                    comm = sp2[4]

                    resource_count[sp2[0]] = int(sp2[1])*int(qt)

                    ress = master_db.objects.get(id = sp2[0])
                    
                    if sp2[0] not in resource_dict:
                        resource_dict[sp2[0]] = float(amnt)*float(qt)
                    else:
                        resource_dict[sp2[0]] += float(amnt)*float(qt)

                    if ress.category not in category_sum:
                        category_sum[ress.category] = float(amnt)*float(qt)
                    else:
                        category_sum[ress.category] += float(amnt)*float(qt)

                    if breakdown_value == "1":
                        sheet.cell(row=ind, column = 2).value = "       " + ress.name
                        sheet.cell(row=ind, column = 2).font = Font(italic=True)
                        sheet.cell(row=ind, column = 3).value = qnt
                        sheet.cell(row=ind, column = 3).font = Font(italic=True)
                        sheet.cell(row=ind, column = 4).value = ress.uom
                        sheet.cell(row=ind, column = 4).font = Font(italic=True)
                        sheet.cell(row=ind, column = 5).value = float(pprc)
                        sheet.cell(row=ind, column = 5).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        sheet.cell(row=ind, column = 5).font = Font(italic=True)
                        sheet.cell(row=ind, column = 6).value = float(amnt)
                        sheet.cell(row=ind, column = 6).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        sheet.cell(row=ind, column = 6).font = Font(italic=True)
                        sheet.cell(row=ind, column = 9).value = comm
                        sheet.cell(row=ind, column = 9).font = Font(italic=True)

                        for k in range(2,7):
                            sheet.cell(row=ind, column = k).fill = res_color
            else:
                total = 0
                for s in sp1:
                    ind += 1
                    sp2 = s.split(",,")
                    qnt = sp2[1]
                    pprc = sp2[2]
                    amnt = sp2[3]
                    comm = sp2[4]

                    resource_count[sp2[0]] = int(sp2[1])*int(qt)

                    ress = master_db.objects.get(id = sp2[0])

                    if sp2[0] not in resource_dict:
                        resource_dict[sp2[0]] = float(amnt)
                    else:
                        resource_dict[sp2[0]] += float(amnt)

                    if ress.category not in category_sum:
                        category_sum[ress.category] = float(amnt)*float(qt)
                    else:
                        category_sum[ress.category] += float(amnt)*float(qt)

                    if breakdown_value == "1":
                        sheet.cell(row=ind, column = 2).value = "       " + ress.name
                        sheet.cell(row=ind, column = 2).font = Font(italic=True)
                        sheet.cell(row=ind, column = 3).value = qnt
                        sheet.cell(row=ind, column = 3).font = Font(italic=True)
                        sheet.cell(row=ind, column = 4).value = ress.uom
                        sheet.cell(row=ind, column = 4).font = Font(italic=True)
                        sheet.cell(row=ind, column = 5).value = float(ress.price)
                        sheet.cell(row=ind, column = 5).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        sheet.cell(row=ind, column = 5).font = Font(italic=True)
                        sheet.cell(row=ind, column = 6).value = ress.price * float(qnt)
                        sheet.cell(row=ind, column = 6).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        sheet.cell(row=ind, column = 6).font = Font(italic=True)
                        sheet.cell(row=ind, column = 9).value = comm
                        sheet.cell(row=ind, column = 9).font = Font(italic=True)

                        for k in range(2,7):
                            sheet.cell(row=ind, column = k).fill = res_color

    sheet.cell(row=1, column=dry_col).value = grand_total_dry
    sheet.cell(row=1, column=dry_col).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    sheet.cell(row=1, column=dry_col).fill = total_color
    sheet.cell(row=1, column=dry_col).font = Font(size=14, bold=True)
    sheet.cell(row=1, column=sell_col).value = grand_total_sell
    sheet.cell(row=1, column=sell_col).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    sheet.cell(row=1, column=sell_col).fill = total_color
    sheet.cell(row=1, column=sell_col).font = Font(size=14, bold=True)
    # for i in range(1,5000):
    #     for j in range(1,10):
    #         sheet.cell(row=i, column = j).border = thin_border

    for i in range(3,20):
        column_number = i
        column = str(chr(64 + column_number))
        sheet.column_dimensions[column].width = 15

    if resource_value == "1":
        sheet2 = wb.create_sheet(title="Resource")

        sheet2.cell(row = 1, column = 4).fill = head_color
        sheet2.cell(row = 1, column = 5).fill = head_color
        sheet2.cell(row = 1, column = 6).fill = head_color
        sheet2.cell(row = 1, column = 7).fill = head_color
        sheet2.cell(row = 1, column = 8).fill = head_color

        column = str(chr(64 + 1))
        sheet2.column_dimensions[column].width = 25

        column = str(chr(64 + 2))
        sheet2.column_dimensions[column].width = 20
        
        column = str(chr(64 + 4))
        sheet2.column_dimensions[column].width = 20

        column = str(chr(64 + 5))
        sheet2.column_dimensions[column].width = 35

        column = str(chr(64 + 6))
        sheet2.column_dimensions[column].width = 15

        column = str(chr(64 + 7))
        sheet2.column_dimensions[column].width = 15

        column = str(chr(64 + 8))
        sheet2.column_dimensions[column].width = 25

        sheet2.cell(row=1, column=4).value = "Code"
        sheet2.cell(row=1, column=4).font = Font(bold=True)
        sheet2.cell(row=1, column=5).value = "Name"
        sheet2.cell(row=1, column=5).font = Font(bold=True)
        sheet2.cell(row=1, column=6).value = "Quantity"
        sheet2.cell(row=1, column=6).font = Font(bold=True)

        sheet2.cell(row=1, column=7).value = "UOM"
        sheet2.cell(row=1, column=7).font = Font(bold=True)
        sheet2.cell(row=1, column=8).value = "Total"
        sheet2.cell(row=1, column=8).font = Font(bold=True)

        rn = 0
        rn+=1
        sheet2.cell(row=rn, column=1).value = "Total Dry Amount (AED)"
        sheet2.cell(row=rn, column=2).value = grand_total_dry
        sheet2.cell(row=rn, column=2).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet2.cell(row = rn, column = 2).font = Font(bold=True)

        rn+=1
        sheet2.cell(row=rn, column=1).value = "Total Selling Amount (AED)"
        sheet2.cell(row=rn, column=2).value = grand_total_sell
        sheet2.cell(row=rn, column=2).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet2.cell(row = rn, column = 2).font = Font(bold=True)

        rn = 1

        for res in resource_dict:
            r = master_db.objects.get(id = res)
            rn+=1
            sheet2.cell(row=rn, column=4).value = r.code
            sheet2.cell(row=rn, column=5).value = r.name
            sheet2.cell(row=rn, column=6).value = resource_count[res]
            sheet2.cell(row=rn, column=6).number_format = numbers.FORMAT_NUMBER
            sheet2.cell(row=rn, column=7).value = r.uom
            sheet2.cell(row=rn, column=8).value = float(resource_dict[res])
            sheet2.cell(row=rn, column=8).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    if category_value == "1":
        sheet3 = wb.create_sheet(title = "Category")

        sheet3.cell(row = 1, column = 4).fill = head_color
        sheet3.cell(row = 1, column = 5).fill = head_color
        sheet3.cell(row = 1, column = 6).fill = head_color

        column = str(chr(64 + 1))
        sheet3.column_dimensions[column].width = 25

        column = str(chr(64 + 2))
        sheet3.column_dimensions[column].width = 20
        
        column = str(chr(64 + 4))
        sheet3.column_dimensions[column].width = 20

        column = str(chr(64 + 5))
        sheet3.column_dimensions[column].width = 25

        column = str(chr(64 + 6))
        sheet3.column_dimensions[column].width = 15

        sheet3.cell(row=1, column=4).value = "Category"
        sheet3.cell(row=1, column=4).font = Font(bold=True)
        sheet3.cell(row=1, column=5).value = "Amount"
        sheet3.cell(row=1, column=5).font = Font(bold=True)
        sheet3.cell(row=1, column=6).value = "Percentange"
        sheet3.cell(row=1, column=6).font = Font(bold=True)

        rn = 0
        rn+=1

        sheet3.cell(row=rn, column=1).value = "Total Dry Amount (AED)"
        sheet3.cell(row=rn, column=2).value = grand_total_dry
        sheet3.cell(row=rn, column=2).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet3.cell(row = rn, column = 2).font = Font(bold=True)

        rn+=1
        sheet3.cell(row=rn, column=1).value = "Total Selling Amount (AED)"
        sheet3.cell(row=rn, column=2).value = grand_total_sell
        sheet3.cell(row=rn, column=2).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sheet3.cell(row = rn, column = 2).font = Font(bold=True)

        rn = 1
        for cat in category_sum:
            rn += 1
            sheet3.cell(row=rn, column=4).value = cat
            sheet3.cell(row=rn, column=5).value = category_sum[cat]
            sheet3.cell(row=rn, column=5).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            sheet3.cell(row=rn, column=6).value = (category_sum[cat]/grand_total_dry)
            sheet3.cell(row=rn, column=6).number_format = numbers.FORMAT_PERCENTAGE_00

    st = "okay"
    try:
        #save_file()
        wb.save("C:/Exports/"+ bo.name +".xlsx")
    except:
        st="error"

    response_data = {}
    response_data['status'] = st
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )
# def save_file():
#     f = asksaveasfile(initialfile = 'Untitled.txt',defaultextension=".txt",filetypes=[("All Files","*.*"),("Text Documents","*.txt")])
# win.mainloop()

def create_resource_row(sheet, ind, mid, boqid):
    master = master_item_edit.objects.filter(master_id=mid, boq_id = boqid)
    tp="edit"
    if len(master) == 0:
        tp="org"
        master = master_item.objects.get(id=mid)
    else:
        master=master[0]

    comps = master.components
    sp1 = comps.split("//")

    for s in sp1:
        ind += 1
        sp2 = s.split(",,")
        sheet.cell(row=ind, column = 2).value = "test"


def get_final_master_price(mid, boqid):
    master = master_item_edit.objects.filter(boq_id = boqid, master_id = mid)
    tp = "edit"
    if len(master) == 0:
        tp="org"
        master = master_item.objects.filter(id = mid)
    else:
        master = master[0]

    t_prc = calculate_master_price_new(master.components, tp)

    return t_prc

def calculate_final_master_price(comps, tp):
    csplit = comps.split("//")
    total = 0
    for cs in csplit:
        split2 = cs.split(",,")
        if tp == "org":
            res = master_db.objects.filter(id = split2[0])
            if len(res) > 0:
                total += float(res[0].price)
        else:
            res = split2[3]
            total += float(res)

    return total

def import_resources(request):
    wb = load_workbook(filename=request.FILES['import_file'].file)
    sheet_obj = wb.active
    row_num = sheet_obj.max_row

    for i in range(2, row_num+1):
        rid = sheet_obj.cell(row=i, column = 1).value
        try:
            res = master_db.objects.get(id = rid)
            res.code = sheet_obj.cell(row=i, column = 2).value
            res.name = sheet_obj.cell(row=i, column = 3).value
            res.price = sheet_obj.cell(row=i, column = 4).value
            res.uom = sheet_obj.cell(row=i, column = 5).value
            res.category = sheet_obj.cell(row=i, column = 6).value
            res.remark = sheet_obj.cell(row=i, column = 7).value
            res.save()
        except:
            res = master_db()
            res.code = sheet_obj.cell(row=i, column = 2).value
            res.name = sheet_obj.cell(row=i, column = 3).value
            res.price = sheet_obj.cell(row=i, column = 4).value
            res.uom = sheet_obj.cell(row=i, column = 5).value
            res.category = sheet_obj.cell(row=i, column = 6).value
            res.remark = sheet_obj.cell(row=i, column = 7).value
            res.save()
    
    return HttpResponseRedirect(reverse('resources'))

@csrf_exempt
def export_resource(request):
    res = master_db.objects.all()
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row = 1, column = 1).value = "id"
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row = 1, column = 2).value = "code"
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row = 1, column = 3).value = "name"
    sheet.cell(row=1, column=3).font = Font(bold=True)
    sheet.cell(row = 1, column = 4).value = "price"
    sheet.cell(row=1, column=4).font = Font(bold=True)
    sheet.cell(row = 1, column = 5).value = "uom"
    sheet.cell(row=1, column=5).font = Font(bold=True)
    sheet.cell(row = 1, column = 6).value = "category"
    sheet.cell(row=1, column=6).font = Font(bold=True)
    sheet.cell(row = 1, column = 7).value = "remark"
    sheet.cell(row=1, column=7).font = Font(bold=True)

    for i in range(len(res)):
        sheet.cell(row = i+2, column = 1).value = res[i].id
        sheet.cell(row = i+2, column = 2).value = res[i].code
        sheet.cell(row = i+2, column = 3).value = res[i].name
        sheet.cell(row = i+2, column = 4).value = res[i].price
        sheet.cell(row = i+2, column = 5).value = res[i].uom
        sheet.cell(row = i+2, column = 6).value = res[i].category
        sheet.cell(row = i+2, column = 7).value = res[i].remark

    #sheet.column_dimensions['A'].hidden= True
    wb.save("Exports/Resources.xlsx")

    response_data = {}
    response_data['status'] = 'done'
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def export_master(request):
    masters = master_item.objects.all()
    with open("Exports/Master_items.pkl", "wb") as f:
        pickle.dump(masters, f)
    
    response_data = {}
    response_data['status'] = 'Export Done'
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

def import_masters(request):
    filename=request.FILES['import_file'].file

    masters = pickle.load(filename)
    for master in masters:
        try:
            m = master_item.objects.get(id = master.id)
            m.code = master.code
            m.name = master.name
            m.components = master.components
            m.remark = master.remark
            m.save()
        except:
            m = master_item()
            m.code = master.code
            m.name = master.name
            m.components = master.components
            m.remark = master.remark
            m.save()

    return HttpResponseRedirect(reverse("masters"))

def all_boqs(request):
    bqs = boq.objects.all()
    return render(request, "manage_boqs.html", {'boqs':bqs})

@csrf_exempt
def update_boq(request):
    data = request.POST.get("data")
    dsp = data.split("/////")
    bq = boq.objects.get(id=int(dsp[0]))

    bq.code = dsp[1]
    bq.name = dsp[2]
    bq.description = dsp[3]
    if dsp[4] == "won":
        bq.won = True
    bq.save()

    response_data = {}
    response_data['data'] = bq.id
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )

@csrf_exempt
def copy_boq(request):
    data = request.POST.get("data")
    dsp = data.split("/////")
    bq = boq.objects.get(id=int(dsp[0]))

    nbq = boq()
    nbq.code = dsp[1]
    nbq.name = dsp[2]
    nbq.description = dsp[3]
    if dsp[4] == "won":
        nbq.won = True
    nbq.data = bq.data
    nbq.factor1 = bq.factor1
    nbq.factor2 = bq.factor2
    nbq.factor3 = bq.factor3
    nbq.factor4 = bq.factor4
    nbq.save()

    response_data = {}
    response_data['data'] = nbq.id
    return HttpResponse(
        json.dumps(response_data),
        content_type = "application/json"
    )